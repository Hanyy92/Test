# Copy from here
"""SOP REPORT - FEISHU BOT VERSION"""
import os
import sys
import json
import glob
import requests
import shutil
import subprocess
import importlib.util
from datetime import datetime

import pandas as pd
from openpyxl import Workbook
from openpyxl.styles import PatternFill, Font, Alignment, Border, Side
from openpyxl.formatting.rule import ColorScaleRule
from openpyxl.utils import get_column_letter

if sys.platform == "win32":
    sys.stdout.reconfigure(encoding="utf-8", errors="replace")

FEISHU_APP_ID = "cli_a9ecf62e46f85cda"
FEISHU_APP_SECRET = "M7ToCZBClLDk9CCqIixFAdLElJqF4mZj"
FEISHU_CHAT_ID = "oc_1ab849cf11a8505ae909eff1928cd052"  # ME CM ONLY
FEISHU_DRIVE_FOLDER_TOKEN = "YOC5fKwB4lqSxIdvUSVc6YmEnLb"
FEISHU_DRIVE_FOLDER_NAME = "SOP Reports"
FEISHU_DEBUG = True
NETLIFY_SITE_URL = os.environ.get("NETLIFY_SITE_URL", "").strip()
HOST_BASE_URL = NETLIFY_SITE_URL
REPO_DIR = os.path.dirname(os.path.abspath(__file__))
OUTPUT_HTML_DIR = os.path.join(REPO_DIR, "output")
PUBLISH_DIR = OUTPUT_HTML_DIR
SOURCE_HTML_DIR = "D:/Daily reports/SOP/output"
HTML_GENERATOR_SCRIPT = "D:/Daily reports/SOP/generate_sop_report_html.py"

# ============================================================================
# REPORT GENERATION CONFIG
# ============================================================================

INPUT_FOLDER = "C:/Users/high tech/Downloads"
OUTPUT_FILE = "D:/Daily reports/SOP/output/Performance_Reports.xlsx"
FILE_PATTERN = "*NEW_SOP*.xlsx"

# ============================================================================
# STYLE DEFINITIONS
# ============================================================================

column_header_fill = PatternFill(start_color="737373", end_color="737373", fill_type="solid")
group_fill = PatternFill(start_color="5A5A5A", end_color="5A5A5A", fill_type="solid")
team_summary_fill = PatternFill(start_color="5A5A5A", end_color="5A5A5A", fill_type="solid")

header_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
normal_font = Font(name="Calibri", size=11)
white_font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")

thin_border = Border(
    left=Side(style="thin"),
    right=Side(style="thin"),
    top=Side(style="thin"),
    bottom=Side(style="thin"),
)

# ============================================================================
# TEAM DEFINITIONS
# ============================================================================

_TEAM_SUFFIX = "小组"
_SUBTOTAL_LABEL = "小计"

eglp_teams = [
    "ME-EGLP-GCC01" + _TEAM_SUFFIX,
    "ME-EGLP01" + _TEAM_SUFFIX,
    "ME-EGLP02" + _TEAM_SUFFIX,
    "ME-EGLP03" + _TEAM_SUFFIX,
    "ME-EGLP04" + _TEAM_SUFFIX,
    "ME-EGLP05" + _TEAM_SUFFIX,
    "ME-EGLP06" + _TEAM_SUFFIX,
    "ME-EGLP07" + _TEAM_SUFFIX,
    "ME-EGLP08" + _TEAM_SUFFIX,
    "ME-EGLP09" + _TEAM_SUFFIX,
    "ME-EGLP10" + _TEAM_SUFFIX,
    "MEAdult-EGLP01" + _TEAM_SUFFIX,
]

jolp_teams = [
    "ME-JOLP02" + _TEAM_SUFFIX,
    "ME-JOLP03" + _TEAM_SUFFIX,
    "ME-JOLP04" + _TEAM_SUFFIX,
    "ME-JOLP06" + _TEAM_SUFFIX,
    "ME-JOLP08" + _TEAM_SUFFIX,
    "MEAdult-JOLP01" + _TEAM_SUFFIX,
]

# SOP types (name, column start index)
sop_info = [
    ("First call", 5),
    ("IUR Feedback", 8),
    ("Service warning - absence", 23),
    ("Service warning - class interruption", 26),
]


class FeishuBot:
    BASE_URL = "https://open.feishu.cn/open-apis"

    def __init__(self, app_id, app_secret):
        self.app_id, self.app_secret = app_id, app_secret
        self._tenant_access_token, self._token_expires_at = None, 0

    def get_tenant_access_token(self):
        if self._tenant_access_token and datetime.now().timestamp() < self._token_expires_at:
            return self._tenant_access_token
        r = requests.post(
            f"{self.BASE_URL}/auth/v3/tenant_access_token/internal/",
            headers={"Content-Type": "application/json"},
            json={"app_id": self.app_id, "app_secret": self.app_secret},
        ).json()
        if r.get("code") != 0:
            raise Exception(f"Auth failed: {r}")
        self._tenant_access_token = r["tenant_access_token"]
        self._token_expires_at = datetime.now().timestamp() + r.get("expire", 7200) - 300
        print("Got token")
        return self._tenant_access_token

    def _request_json(self, method, url, **kwargs):
        r = requests.request(method, url, **kwargs)
        try:
            return r.json()
        except Exception:
            if FEISHU_DEBUG:
                print(f"Non-JSON response from {url}: {r.status_code} {r.text[:2000]}")
            raise

    def upload_image(self, p):
        with open(p, "rb") as f:
            r = self._request_json(
                "POST",
                f"{self.BASE_URL}/im/v1/images",
                headers={"Authorization": f"Bearer {self.get_tenant_access_token()}"},
                files={"image": (os.path.basename(p), f, "image/png")},
                data={"image_type": "message"},
            )
        if r.get("code") != 0:
            raise Exception(f"Upload failed: {r}")
        print(f"Uploaded: {os.path.basename(p)}")
        return r["data"]["image_key"]

    def list_drive_files(self, folder_token):
        r = self._request_json(
            "GET",
            f"{self.BASE_URL}/drive/v1/files",
            headers={"Authorization": f"Bearer {self.get_tenant_access_token()}"},
            params={"folder_token": folder_token, "page_size": 200},
        )
        if r.get("code") != 0:
            raise Exception(f"List drive files failed: {r}")
        return r.get("data", {}).get("files", [])

    def delete_drive_file(self, file_token):
        r = self._request_json(
            "DELETE",
            f"{self.BASE_URL}/drive/v1/files/{file_token}",
            headers={"Authorization": f"Bearer {self.get_tenant_access_token()}"},
            params={"type": "file"},
        )
        if r.get("code") != 0:
            raise Exception(f"Delete drive file failed: {r}")

    def upload_drive_file(self, p, folder_token):
        size = os.path.getsize(p)
        with open(p, "rb") as f:
            r = self._request_json(
                "POST",
                f"{self.BASE_URL}/drive/v1/files/upload_all",
                headers={"Authorization": f"Bearer {self.get_tenant_access_token()}"},
                data={
                    "file_name": os.path.basename(p),
                    "parent_type": "explorer",
                    "parent_node": folder_token,
                    "size": str(size),
                    "file_type": "stream",
                },
                files={"file": (os.path.basename(p), f)},
            )
        if r.get("code") != 0:
            raise Exception(f"Drive upload failed: {r}")
        return r["data"]["file_token"]

    def create_drive_share_link(self, file_token):
        r = self._request_json(
            "POST",
            f"{self.BASE_URL}/drive/v1/permissions/create",
            headers={
                "Authorization": f"Bearer {self.get_tenant_access_token()}",
                "Content-Type": "application/json",
            },
            json={"type": "share", "token": file_token, "perm": "view"},
        )
        if r.get("code") != 0:
            raise Exception(f"Create share link failed: {r}")
        data = r.get("data", {})
        return data.get("url") or data.get("share_url") or data.get("link_url")

    def create_drive_folder(self, parent_node, folder_name):
        r = self._request_json(
            "POST",
            f"{self.BASE_URL}/drive/v1/files/create_folder",
            headers={
                "Authorization": f"Bearer {self.get_tenant_access_token()}",
                "Content-Type": "application/json",
            },
            json={"name": folder_name, "parent_type": "explorer", "parent_node": parent_node},
        )
        if r.get("code") != 0:
            raise Exception(f"Create drive folder failed: {r}")
        data = r.get("data", {})
        return data.get("token") or data.get("file_token") or data.get("folder_token")

    def ensure_drive_folder(self, folder_token, folder_name):
        if folder_token and folder_token.upper() != "AUTO":
            return folder_token
        try:
            items = self.list_drive_files("root")
            for item in items:
                name = item.get("name")
                if name == folder_name and (item.get("type") == "folder" or item.get("file_type") == "folder"):
                    return item.get("token") or item.get("file_token") or item.get("folder_token")
        except Exception as e:
            if FEISHU_DEBUG:
                print(f"Drive root list failed: {e}")
        return self.create_drive_folder("root", folder_name)

    def get_bot_chats(self):
        r = self._request_json(
            "GET",
            f"{self.BASE_URL}/im/v1/chats",
            headers={"Authorization": f"Bearer {self.get_tenant_access_token()}"},
        )
        if r.get("code") != 0:
            raise Exception(f"Get chats failed: {r}")
        return r.get("data", {}).get("items", [])

    def send_interactive_card(self, cid, card):
        r = self._request_json(
            "POST",
            f"{self.BASE_URL}/im/v1/messages",
            headers={
                "Authorization": f"Bearer {self.get_tenant_access_token()}",
                "Content-Type": "application/json",
            },
            params={"receive_id_type": "chat_id"},
            json={
                "receive_id": cid,
                "msg_type": "interactive",
                "content": json.dumps(card),
            },
        )
        if r.get("code") != 0:
            raise Exception(f"Send card failed: {r}")

    def build_sop_reports_card(self, links):
        def preview_action(link_url, label):
            plain_label = label.replace("**", "")
            return {
                "tag": "action",
                "actions": [
                    {
                        "tag": "button",
                        "text": {"tag": "plain_text", "content": f"View {plain_label}"},
                        "type": "primary",
                        "multi_url": {
                            "url": link_url,
                            "pc_url": link_url,
                            "ios_url": link_url,
                            "android_url": link_url,
                        },
                    }
                ],
            }

        els = [
            {
                "tag": "div",
                "text": {
                    "tag": "lark_md",
                    "content": f"**Report Time:** {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}",
                },
            },
            {"tag": "hr"},
        ]
        for url, lbl in links:
            els.extend(
                [
                    {"tag": "div", "text": {"tag": "lark_md", "content": lbl}},
                    preview_action(url, lbl),
                    {"tag": "hr"},
                ]
            )
        if els and els[-1].get("tag") == "hr":
            els.pop()
        return {
            "config": {"wide_screen_mode": True},
            "header": {"title": {"tag": "plain_text", "content": "January SOP Reports"}, "template": "blue"},
            "elements": els,
        }


def find_newest_file(folder_path, pattern="*.xlsx"):
    search_path = os.path.join(folder_path, pattern)
    files = glob.glob(search_path)
    files = [f for f in files if not os.path.basename(f).startswith("~$")]
    if not files:
        return None
    return max(files, key=os.path.getmtime)


def prepare_group_data(df, teams_list):
    group_data = df[df["org_name"].isin(teams_list)].copy()
    summaries = group_data[group_data["user_account"] == _SUBTOTAL_LABEL].copy()
    agents = group_data[group_data["user_account"] != _SUBTOTAL_LABEL].copy()
    return agents, summaries


def create_separate_team_sheet(ws, agents_df, summaries_df, title, team_order, sort_column):
    teams_dict = {}
    for team_name in team_order:
        team_summary_rows = summaries_df[summaries_df["org_name"] == team_name]
        if len(team_summary_rows) > 0:
            team_summary = team_summary_rows.iloc[0]
            team_agents = agents_df[agents_df["org_name"] == team_name].sort_values(sort_column, ascending=False)
            teams_dict[team_name] = {"agents": team_agents, "summary": team_summary}

    ws.merge_cells("A1:N1")
    title_cell = ws["A1"]
    title_cell.value = title
    title_cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    title_cell.fill = column_header_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:A3")
    cell = ws["A2"]
    cell.value = "Group"
    cell.font = header_font
    cell.fill = column_header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

    ws.merge_cells("B2:B3")
    cell = ws["B2"]
    cell.value = "Agent"
    cell.font = header_font
    cell.fill = column_header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

    col_idx = 3
    for sop_name, _ in sop_info:
        start_col = get_column_letter(col_idx)
        end_col = get_column_letter(col_idx + 2)
        ws.merge_cells(f"{start_col}2:{end_col}2")
        cell = ws.cell(row=2, column=col_idx)
        cell.value = sop_name
        cell.font = header_font
        cell.fill = column_header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

        for c in range(col_idx, col_idx + 3):
            ws.cell(row=2, column=c).border = thin_border

        col_idx += 3

    ws.row_dimensions[2].height = 30

    col_idx = 3
    for sop_name, _ in sop_info:
        cell = ws.cell(row=3, column=col_idx)
        cell.value = "Total SOPs"
        cell.font = header_font
        cell.fill = column_header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        col_idx += 1

        cell = ws.cell(row=3, column=col_idx)
        cell.value = "Finished"
        cell.font = header_font
        cell.fill = column_header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        col_idx += 1

        cell = ws.cell(row=3, column=col_idx)
        cell.value = "Dropped"
        cell.font = header_font
        cell.fill = column_header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        col_idx += 1

    ws.row_dimensions[3].height = 40

    current_row = 4
    ontime_cols = [4 + i * 3 for i in range(len(sop_info))]
    first_team = True

    for team_name in team_order:
        if team_name not in teams_dict:
            continue
        team_data = teams_dict[team_name]
        agents = team_data["agents"]
        summary = team_data["summary"]
        team_start_row = current_row

        if not first_team:
            ws.merge_cells(f"A{current_row}:A{current_row + 1}")
            cell = ws.cell(row=current_row, column=1)
            cell.value = "Group"
            cell.font = header_font
            cell.fill = column_header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

            ws.merge_cells(f"B{current_row}:B{current_row + 1}")
            cell = ws.cell(row=current_row, column=2)
            cell.value = "Agent"
            cell.font = header_font
            cell.fill = column_header_fill
            cell.alignment = Alignment(horizontal="center", vertical="center")
            cell.border = thin_border

            col_idx = 3
            for sop_name, _ in sop_info:
                start_col = get_column_letter(col_idx)
                end_col = get_column_letter(col_idx + 2)
                ws.merge_cells(f"{start_col}{current_row}:{end_col}{current_row}")
                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = sop_name
                cell.font = header_font
                cell.fill = column_header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center")
                cell.border = thin_border

                for c in range(col_idx, col_idx + 3):
                    ws.cell(row=current_row, column=c).border = thin_border

                col_idx += 3

            ws.row_dimensions[current_row].height = 25
            current_row += 1

            col_idx = 3
            for sop_name, _ in sop_info:
                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = "Total SOPs"
                cell.font = header_font
                cell.fill = column_header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border
                col_idx += 1

                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = "Finished"
                cell.font = header_font
                cell.fill = column_header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border
                col_idx += 1

                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = "Dropped"
                cell.font = header_font
                cell.fill = column_header_fill
                cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
                cell.border = thin_border
                col_idx += 1

            ws.row_dimensions[current_row].height = 30
            current_row += 1

            team_start_row = current_row

        first_team = False

        for _, agent_row in agents.iterrows():
            ws.cell(row=current_row, column=2).value = agent_row["user_account"]
            ws.cell(row=current_row, column=2).font = normal_font
            ws.cell(row=current_row, column=2).border = thin_border
            ws.cell(row=current_row, column=2).alignment = Alignment(horizontal="center", vertical="center")

            col_idx = 3
            for _, df_col_start in sop_info:
                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = agent_row.iloc[df_col_start]
                cell.font = normal_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                col_idx += 1

                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = agent_row.iloc[df_col_start + 1]
                cell.number_format = "0.0%"
                cell.font = normal_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                col_idx += 1

                cell = ws.cell(row=current_row, column=col_idx)
                cell.value = agent_row.iloc[df_col_start + 2]
                cell.number_format = "0.0%"
                cell.font = normal_font
                cell.border = thin_border
                cell.alignment = Alignment(horizontal="center", vertical="center")
                col_idx += 1

            ws.row_dimensions[current_row].height = 40
            current_row += 1

        team_end_agents = current_row - 1

        ws.cell(row=current_row, column=2).value = team_name
        ws.cell(row=current_row, column=2).font = white_font
        ws.cell(row=current_row, column=2).fill = team_summary_fill
        ws.cell(row=current_row, column=2).border = thin_border
        ws.cell(row=current_row, column=2).alignment = Alignment(horizontal="center", vertical="center")

        col_idx = 3
        for _, df_col_start in sop_info:
            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = summary.iloc[df_col_start]
            cell.font = white_font
            cell.fill = team_summary_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            col_idx += 1

            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = summary.iloc[df_col_start + 1]
            cell.number_format = "0.0%"
            cell.font = white_font
            cell.fill = team_summary_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            col_idx += 1

            cell = ws.cell(row=current_row, column=col_idx)
            cell.value = summary.iloc[df_col_start + 2]
            cell.number_format = "0.0%"
            cell.font = white_font
            cell.fill = team_summary_fill
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            col_idx += 1

        ws.row_dimensions[current_row].height = 40
        current_row += 1

        team_end_row = current_row - 1
        ws.merge_cells(f"A{team_start_row}:A{team_end_row}")
        group_cell = ws.cell(row=team_start_row, column=1)
        group_cell.value = team_name
        group_cell.font = white_font
        group_cell.fill = group_fill
        group_cell.alignment = Alignment(horizontal="center", vertical="center")
        group_cell.border = thin_border

        if team_end_agents >= team_start_row:
            for col_num in ontime_cols:
                col_letter = get_column_letter(col_num)
                color_scale = ColorScaleRule(
                    start_type="min",
                    start_color="F8696B",
                    mid_type="percentile",
                    mid_value=50,
                    mid_color="FFEB84",
                    end_type="max",
                    end_color="63BE7B",
                )
                ws.conditional_formatting.add(
                    f"{col_letter}{team_start_row}:{col_letter}{team_end_agents}",
                    color_scale,
                )

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 22
    for i in range(3, 3 + len(sop_info) * 3):
        ws.column_dimensions[get_column_letter(i)].width = 13


def create_ranking_sheet(ws, agents_df, title, sort_column):
    agents_sorted = agents_df.sort_values(sort_column, ascending=False).reset_index(drop=True)

    ws.merge_cells("A1:N1")
    title_cell = ws["A1"]
    title_cell.value = title
    title_cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    title_cell.fill = column_header_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:A3")
    cell = ws["A2"]
    cell.value = "Group"
    cell.font = header_font
    cell.fill = column_header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

    ws.merge_cells("B2:B3")
    cell = ws["B2"]
    cell.value = "Agent"
    cell.font = header_font
    cell.fill = column_header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

    col_idx = 3
    for sop_name, _ in sop_info:
        start_col = get_column_letter(col_idx)
        end_col = get_column_letter(col_idx + 2)
        ws.merge_cells(f"{start_col}2:{end_col}2")
        cell = ws.cell(row=2, column=col_idx)
        cell.value = sop_name
        cell.font = header_font
        cell.fill = column_header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

        for c in range(col_idx, col_idx + 3):
            ws.cell(row=2, column=c).border = thin_border

        col_idx += 3

    ws.row_dimensions[2].height = 30

    col_idx = 3
    for sop_name, _ in sop_info:
        cell = ws.cell(row=3, column=col_idx)
        cell.value = "Total SOPs"
        cell.font = header_font
        cell.fill = column_header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        col_idx += 1

        cell = ws.cell(row=3, column=col_idx)
        cell.value = "Finished"
        cell.font = header_font
        cell.fill = column_header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        col_idx += 1

        cell = ws.cell(row=3, column=col_idx)
        cell.value = "Dropped"
        cell.font = header_font
        cell.fill = column_header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        col_idx += 1

    ws.row_dimensions[3].height = 40

    ontime_cols = [4 + i * 3 for i in range(len(sop_info))]

    for idx, agent_row in agents_sorted.iterrows():
        row_num = idx + 4

        cell = ws.cell(row=row_num, column=1)
        cell.value = agent_row["org_name"]
        cell.font = white_font
        cell.fill = group_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

        cell = ws.cell(row=row_num, column=2)
        cell.value = agent_row["user_account"]
        cell.font = normal_font
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

        col_idx = 3
        for _, df_col_start in sop_info:
            cell = ws.cell(row=row_num, column=col_idx)
            cell.value = agent_row.iloc[df_col_start]
            cell.font = normal_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            col_idx += 1

            cell = ws.cell(row=row_num, column=col_idx)
            cell.value = agent_row.iloc[df_col_start + 1]
            cell.number_format = "0.0%"
            cell.font = normal_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            col_idx += 1

            cell = ws.cell(row=row_num, column=col_idx)
            cell.value = agent_row.iloc[df_col_start + 2]
            cell.number_format = "0.0%"
            cell.font = normal_font
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            col_idx += 1

        ws.row_dimensions[row_num].height = 40

    if len(agents_sorted) > 0:
        for col_num in ontime_cols:
            col_letter = get_column_letter(col_num)
            color_scale = ColorScaleRule(
                start_type="min",
                start_color="F8696B",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFEB84",
                end_type="max",
                end_color="63BE7B",
            )
            ws.conditional_formatting.add(
                f"{col_letter}4:{col_letter}{len(agents_sorted) + 3}",
                color_scale,
            )

    ws.column_dimensions["A"].width = 18
    ws.column_dimensions["B"].width = 22
    for i in range(3, 3 + len(sop_info) * 3):
        ws.column_dimensions[get_column_letter(i)].width = 13


def create_teams_totals_sheet(ws, summaries_df, title, sort_column):
    summaries_sorted = summaries_df.sort_values(sort_column, ascending=False).reset_index(drop=True)

    ws.merge_cells("A1:M1")
    title_cell = ws["A1"]
    title_cell.value = title
    title_cell.font = Font(name="Calibri", size=14, bold=True, color="FFFFFF")
    title_cell.fill = column_header_fill
    title_cell.alignment = Alignment(horizontal="center", vertical="center")
    ws.row_dimensions[1].height = 40

    ws.merge_cells("A2:A3")
    cell = ws["A2"]
    cell.value = "Group"
    cell.font = header_font
    cell.fill = column_header_fill
    cell.alignment = Alignment(horizontal="center", vertical="center")
    cell.border = thin_border

    col_idx = 2
    for sop_name, _ in sop_info:
        start_col = get_column_letter(col_idx)
        end_col = get_column_letter(col_idx + 2)
        ws.merge_cells(f"{start_col}2:{end_col}2")
        cell = ws.cell(row=2, column=col_idx)
        cell.value = sop_name
        cell.font = header_font
        cell.fill = column_header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center")
        cell.border = thin_border

        for c in range(col_idx, col_idx + 3):
            ws.cell(row=2, column=c).border = thin_border

        col_idx += 3

    ws.row_dimensions[2].height = 30

    col_idx = 2
    for sop_name, _ in sop_info:
        cell = ws.cell(row=3, column=col_idx)
        cell.value = "Total SOPs"
        cell.font = header_font
        cell.fill = column_header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        col_idx += 1

        cell = ws.cell(row=3, column=col_idx)
        cell.value = "Finished"
        cell.font = header_font
        cell.fill = column_header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        col_idx += 1

        cell = ws.cell(row=3, column=col_idx)
        cell.value = "Dropped"
        cell.font = header_font
        cell.fill = column_header_fill
        cell.alignment = Alignment(horizontal="center", vertical="center", wrap_text=True)
        cell.border = thin_border
        col_idx += 1

    ws.row_dimensions[3].height = 40

    ontime_cols = [3 + i * 3 for i in range(len(sop_info))]

    for idx, team_row in summaries_sorted.iterrows():
        row_num = idx + 4

        cell = ws.cell(row=row_num, column=1)
        cell.value = team_row["org_name"]
        cell.font = white_font
        cell.fill = group_fill
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")

        col_idx = 2
        for _, df_col_start in sop_info:
            cell = ws.cell(row=row_num, column=col_idx)
            cell.value = team_row.iloc[df_col_start]
            cell.font = Font(name="Calibri", size=11, bold=True)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            col_idx += 1

            cell = ws.cell(row=row_num, column=col_idx)
            cell.value = team_row.iloc[df_col_start + 1]
            cell.number_format = "0.0%"
            cell.font = Font(name="Calibri", size=11, bold=True)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            col_idx += 1

            cell = ws.cell(row=row_num, column=col_idx)
            cell.value = team_row.iloc[df_col_start + 2]
            cell.number_format = "0.0%"
            cell.font = Font(name="Calibri", size=11, bold=True)
            cell.border = thin_border
            cell.alignment = Alignment(horizontal="center", vertical="center")
            col_idx += 1

        ws.row_dimensions[row_num].height = 40

    total_row_num = len(summaries_sorted) + 4

    cell = ws.cell(row=total_row_num, column=1)
    cell.value = "TOTAL"
    cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
    cell.fill = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
    cell.border = thin_border
    cell.alignment = Alignment(horizontal="center", vertical="center")

    col_idx = 2
    for _, _ in sop_info:
        total_sops = 0
        for row_num in range(4, 4 + len(summaries_sorted)):
            cell_value = ws.cell(row=row_num, column=col_idx).value
            if cell_value is not None:
                total_sops += cell_value

        cell = ws.cell(row=total_row_num, column=col_idx)
        cell.value = total_sops
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        col_idx += 1

        sum_percent = 0
        count_teams = 0
        for row_num in range(4, 4 + len(summaries_sorted)):
            percent = ws.cell(row=row_num, column=col_idx).value
            if percent is not None:
                sum_percent += percent
                count_teams += 1

        avg_finished = sum_percent / count_teams if count_teams > 0 else 0
        cell = ws.cell(row=total_row_num, column=col_idx)
        cell.value = avg_finished
        cell.number_format = "0.0%"
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        col_idx += 1

        sum_percent = 0
        count_teams = 0
        for row_num in range(4, 4 + len(summaries_sorted)):
            percent = ws.cell(row=row_num, column=col_idx).value
            if percent is not None:
                sum_percent += percent
                count_teams += 1

        avg_dropped = sum_percent / count_teams if count_teams > 0 else 0
        cell = ws.cell(row=total_row_num, column=col_idx)
        cell.value = avg_dropped
        cell.number_format = "0.0%"
        cell.font = Font(name="Calibri", size=11, bold=True, color="FFFFFF")
        cell.fill = PatternFill(start_color="404040", end_color="404040", fill_type="solid")
        cell.border = thin_border
        cell.alignment = Alignment(horizontal="center", vertical="center")
        col_idx += 1

    ws.row_dimensions[total_row_num].height = 40

    if len(summaries_sorted) > 0:
        for col_num in ontime_cols:
            col_letter = get_column_letter(col_num)
            color_scale = ColorScaleRule(
                start_type="min",
                start_color="F8696B",
                mid_type="percentile",
                mid_value=50,
                mid_color="FFEB84",
                end_type="max",
                end_color="63BE7B",
            )
            ws.conditional_formatting.add(
                f"{col_letter}4:{col_letter}{len(summaries_sorted) + 3}",
                color_scale,
            )

    ws.column_dimensions["A"].width = 18
    for i in range(2, 2 + len(sop_info) * 3):
        ws.column_dimensions[get_column_letter(i)].width = 13


def generate_sop_report():
    print("=" * 80)
    print("DAILY SOP REPORTS GENERATOR")
    print("=" * 80)
    print(f"\nRun time: {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}")

    input_file = find_newest_file(INPUT_FOLDER, FILE_PATTERN)
    if input_file is None:
        print(f"\nERROR: No Excel files found in {INPUT_FOLDER}")
        return None

    file_modified_time = datetime.fromtimestamp(os.path.getmtime(input_file))
    file_name = os.path.basename(input_file)
    print("\nFound newest file:")
    print(f"  Name: {file_name}")
    print(f"  Modified: {file_modified_time.strftime('%Y-%m-%d %H:%M:%S')}")
    print(f"  Path: {input_file}")

    file = pd.read_excel(input_file)
    data = file.iloc[1:].reset_index(drop=True).copy()

    for col in data.columns[2:]:
        data[col] = data[col].replace("-", 0)
        data[col] = data[col].replace("", 0)
        data[col] = pd.to_numeric(data[col], errors="coerce").fillna(0)

    data.rename(columns={data.columns[0]: "org_name", data.columns[1]: "user_account"}, inplace=True)

    excluded_agents = [
        "JOSS-hadeelrabie",
        "JOLP-naserabdallah",
        "JOSS-masaadeh",
        "JOSS-marwatarawneh",
        "JOSS-nourdaabes",
    ]
    is_subtotal = data["user_account"] == _SUBTOTAL_LABEL
    is_excluded_agent = data["user_account"].isin(excluded_agents)
    data = data[~(is_excluded_agent & ~is_subtotal)].copy()

    sort_column = data.columns[6]

    eglp_agents, eglp_summaries = prepare_group_data(data, eglp_teams)
    jolp_agents, jolp_summaries = prepare_group_data(data, jolp_teams)

    wb = Workbook()
    wb.remove(wb.active)

    ws1 = wb.create_sheet("EGLP Teams")
    create_separate_team_sheet(
        ws1,
        eglp_agents,
        eglp_summaries,
        "SOP Report - EGLP Teams",
        eglp_teams,
        sort_column,
    )

    ws2 = wb.create_sheet("EGLP Ranking")
    create_ranking_sheet(ws2, eglp_agents, "SOP Report - EGLP Ranking", sort_column)

    ws3 = wb.create_sheet("EGLP Teams Totals")
    create_teams_totals_sheet(
        ws3,
        eglp_summaries,
        "SOP Report - EGLP Teams Totals",
        sort_column,
    )

    ws4 = wb.create_sheet("JOLP Teams")
    create_separate_team_sheet(
        ws4,
        jolp_agents,
        jolp_summaries,
        "SOP Report - JOLP Teams",
        jolp_teams,
        sort_column,
    )

    ws5 = wb.create_sheet("JOLP Ranking")
    create_ranking_sheet(ws5, jolp_agents, "SOP Report - JOLP Ranking", sort_column)

    ws6 = wb.create_sheet("JOLP Teams Totals")
    create_teams_totals_sheet(
        ws6,
        jolp_summaries,
        "SOP Report - JOLP Teams Totals",
        sort_column,
    )

    os.makedirs(os.path.dirname(OUTPUT_FILE), exist_ok=True)
    wb.save(OUTPUT_FILE)
    print(f"\nReports generated: {OUTPUT_FILE}")
    return OUTPUT_FILE


def send_reports_to_feishu(output_dir, chat_id=None):
    bot = FeishuBot(FEISHU_APP_ID, FEISHU_APP_SECRET)
    if not chat_id:
        chats = bot.get_bot_chats()
        if not chats:
            print("ERROR: No chats")
            return False
        chat_id = chats[0].get("chat_id")
    if not publish_reports_to_repo(output_dir):
        print("ERROR: Failed to publish reports to repo.")
        return False
    links = get_report_page_links()
    if not links:
        print("ERROR: No report page links")
        return False
    try:
        card = bot.build_sop_reports_card(links)
        bot.send_interactive_card(chat_id, card)
        print("Sent card with all images!")
        return True
    except Exception as e:
        print(f"ERROR: {e}")
        import traceback
        traceback.print_exc()
        return False


def sync_report_pages_to_repo(output_dir, publish_dir):
    os.makedirs(publish_dir, exist_ok=True)
    copied = 0
    missing = []
    same_dir = os.path.abspath(output_dir) == os.path.abspath(publish_dir)
    for filename, _, _ in get_report_pages():
        src = os.path.join(output_dir, filename)
        if not os.path.exists(src):
            missing.append(filename)
            continue
        if not same_dir:
            dst = os.path.join(publish_dir, filename)
            shutil.copy2(src, dst)
            copied += 1
        if filename == "Performance_Reports_overview.html":
            index_path = os.path.join(publish_dir, "index.html")
            if os.path.abspath(src) != os.path.abspath(index_path):
                shutil.copy2(src, index_path)
    if missing:
        print(f"WARNING: Missing HTML pages: {', '.join(missing)}")
    if copied == 0 and not same_dir:
        print("ERROR: No HTML pages copied to publish dir.")
        return False
    if same_dir:
        print(f"Using HTML output dir as publish dir: {publish_dir}")
    else:
        print(f"Copied {copied} HTML pages to publish dir: {publish_dir}")
    return True


def git_has_changes(repo_dir):
    r = subprocess.run(
        ["git", "status", "--porcelain"],
        cwd=repo_dir,
        capture_output=True,
        text=True,
        check=False,
    )
    if r.returncode != 0:
        print(f"Git status failed: {r.stderr.strip()}")
        return False, False
    return True, bool(r.stdout.strip())


def git_commit_and_push(repo_dir, message):
    r = subprocess.run(["git", "add", "-A"], cwd=repo_dir, check=False)
    if r.returncode != 0:
        print("Git add failed.")
        return False
    ok, has_changes = git_has_changes(repo_dir)
    if not ok:
        return False
    if not has_changes:
        print("No changes to commit.")
        return True
    r = subprocess.run(["git", "commit", "-m", message], cwd=repo_dir, check=False)
    if r.returncode != 0:
        print("Git commit failed.")
        return False
    r = subprocess.run(["git", "push"], cwd=repo_dir, check=False)
    if r.returncode != 0:
        print("Git push failed.")
        return False
    return True


def publish_reports_to_repo(output_dir):
    if not sync_report_pages_to_repo(output_dir, PUBLISH_DIR):
        return False
    msg = f"Update SOP HTML dashboards {datetime.now().strftime('%Y-%m-%d %H:%M:%S')}"
    return git_commit_and_push(REPO_DIR, msg)


def get_report_pages():
    return [
        ("Performance_Reports_overview.html", "overview", "**ME Total**"),
        ("Performance_Reports_eglp_teams.html", "eglp-teams", "**EG Teams**"),
        ("Performance_Reports_eglp_ranking.html", "eglp-ranking", "**EGLP Ranking**"),
        ("Performance_Reports_eglp_totals.html", "eglp-totals", "**EGLP Totals**"),
        ("Performance_Reports_jolp_teams.html", "jolp-teams", "**JOLP Teams**"),
        ("Performance_Reports_jolp_ranking.html", "jolp-ranking", "**JOLP Ranking**"),
        ("Performance_Reports_jolp_totals.html", "jolp-totals", "**JOLP Totals**"),
    ]


def generate_html_reports_with_original_script():
    if not os.path.exists(HTML_GENERATOR_SCRIPT):
        print(f"ERROR: HTML generator script not found: {HTML_GENERATOR_SCRIPT}")
        return False
    try:
        spec = importlib.util.spec_from_file_location("sop_html_generator", HTML_GENERATOR_SCRIPT)
        module = importlib.util.module_from_spec(spec)
        spec.loader.exec_module(module)
        module.SKIP_UPLOAD_AND_SEND = True
        if NETLIFY_SITE_URL:
            module.HOST_BASE_URL = NETLIFY_SITE_URL
        module.main()
        return True
    except Exception as e:
        print(f"ERROR: HTML generator failed: {e}")
        import traceback
        traceback.print_exc()
        return False


def copy_html_reports(source_dir, html_output_dir):
    os.makedirs(html_output_dir, exist_ok=True)
    copied = 0
    missing = []
    for filename, _, _ in get_report_pages():
        src = os.path.join(source_dir, filename)
        if not os.path.exists(src):
            missing.append(filename)
            continue
        dst = os.path.join(html_output_dir, filename)
        shutil.copy2(src, dst)
        copied += 1
        if filename == "Performance_Reports_overview.html":
            shutil.copy2(src, os.path.join(html_output_dir, "index.html"))
    if missing:
        print(f"WARNING: Missing HTML pages in source dir: {', '.join(missing)}")
    if copied == 0:
        print("ERROR: No HTML reports copied.")
        return False
    print(f"Copied {copied} HTML reports from {source_dir} to {html_output_dir}")
    return True


def get_report_page_links():
    links = []
    if not HOST_BASE_URL:
        print("WARNING: NETLIFY_SITE_URL not set; using relative links.")
    for filename, _, label in get_report_pages():
        if HOST_BASE_URL:
            url = f"{HOST_BASE_URL.rstrip('/')}/{filename}"
        else:
            url = filename
        links.append((url, label))
    return links


def main():
    o = generate_sop_report()
    if o is None:
        return
    if not generate_html_reports_with_original_script():
        return
    if not copy_html_reports(SOURCE_HTML_DIR, OUTPUT_HTML_DIR):
        return
    send_reports_to_feishu(OUTPUT_HTML_DIR, FEISHU_CHAT_ID)


if __name__ == "__main__":
    main()
